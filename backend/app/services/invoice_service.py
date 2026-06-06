from fastapi import HTTPException, status, UploadFile
from sqlalchemy.ext.asyncio import AsyncSession
from typing import Optional, List
import math

from app.repositories.invoice_repository import InvoiceRepository, FileRepository
from app.repositories.audit_repository import AuditRepository
from app.schemas.schemas import (
    InvoiceCreate, InvoiceUpdate, InvoiceResponse, InvoiceListResponse, InvoiceFilters, UploadedFileResponse
)

ALLOWED_CONTENT_TYPES = {
    "application/pdf",
    "image/jpeg",
    "image/jpg",
    "image/png",
}
MAX_FILE_SIZE = 20 * 1024 * 1024  # 20MB


class InvoiceService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.invoice_repo = InvoiceRepository(db)
        self.file_repo = FileRepository(db)
        self.audit_repo = AuditRepository(db)

    async def create_invoice(self, data: InvoiceCreate, user_id: str, ip: str) -> InvoiceResponse:
        invoice = await self.invoice_repo.create(
            bank_name=data.bank_name,
            invoice_date=data.invoice_date,
            loan_requested_amount=data.loan_requested_amount,
            loan_sanctioned_amount=data.loan_sanctioned_amount,
        )
        await self.audit_repo.log(
            action="invoice_created",
            user_id=user_id,
            resource_type="invoice",
            resource_id=invoice.id,
            ip_address=ip,
            details=f"Invoice {invoice.invoice_number} created",
        )
        return InvoiceResponse.model_validate(invoice)

    async def update_invoice(self, invoice_id: str, data: InvoiceUpdate, user_id: str, ip: str) -> InvoiceResponse:
        invoice = await self.invoice_repo.get_by_id(invoice_id)
        if not invoice:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Invoice not found")

        update_data = {k: v for k, v in data.model_dump(exclude_unset=True).items()}
        if not update_data:
            return InvoiceResponse.model_validate(invoice)

        updated = await self.invoice_repo.update(invoice_id, **update_data)
        await self.audit_repo.log(
            action="invoice_updated",
            user_id=user_id,
            resource_type="invoice",
            resource_id=invoice_id,
            ip_address=ip,
        )
        return InvoiceResponse.model_validate(updated)

    async def delete_invoice(self, invoice_id: str, user_id: str, ip: str):
        invoice = await self.invoice_repo.get_by_id(invoice_id)
        if not invoice:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Invoice not found")

        await self.invoice_repo.delete(invoice_id)
        await self.audit_repo.log(
            action="invoice_deleted",
            user_id=user_id,
            resource_type="invoice",
            resource_id=invoice_id,
            ip_address=ip,
            details=f"Invoice {invoice.invoice_number} deleted",
        )

    async def get_invoice(self, invoice_id: str) -> InvoiceResponse:
        invoice = await self.invoice_repo.get_by_id(invoice_id)
        if not invoice:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Invoice not found")
        return InvoiceResponse.model_validate(invoice)

    async def list_invoices(self, filters: InvoiceFilters) -> InvoiceListResponse:
        items, total = await self.invoice_repo.list_with_filters(filters)
        total_pages = math.ceil(total / filters.page_size) if total > 0 else 1
        return InvoiceListResponse(
            items=[InvoiceResponse.model_validate(i) for i in items],
            total=total,
            page=filters.page,
            page_size=filters.page_size,
            total_pages=total_pages,
        )

    async def upload_file(self, invoice_id: str, file: UploadFile, user_id: str, ip: str):
        invoice = await self.invoice_repo.get_by_id(invoice_id)
        if not invoice:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Invoice not found")

        if file.content_type not in ALLOWED_CONTENT_TYPES:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File type not allowed. Supported: PDF, JPG, JPEG, PNG",
            )

        file_data = await file.read()
        if len(file_data) > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File size exceeds 20MB limit",
            )

        existing_file = await self.file_repo.get_by_invoice_id(invoice_id)
        if existing_file:
            updated = await self.file_repo.update(
                file_id=existing_file.id,
                filename=file.filename,
                content_type=file.content_type,
                file_size=len(file_data),
                file_data=file_data,
            )
            await self.audit_repo.log(
                action="file_replaced",
                user_id=user_id,
                resource_type="invoice",
                resource_id=invoice_id,
                ip_address=ip,
                details=f"File replaced: {file.filename}",
            )
            return UploadedFileResponse.model_validate(updated)
        else:
            new_file = await self.file_repo.create(
                invoice_id=invoice_id,
                filename=file.filename,
                content_type=file.content_type,
                file_size=len(file_data),
                file_data=file_data,
            )
            await self.audit_repo.log(
                action="file_uploaded",
                user_id=user_id,
                resource_type="invoice",
                resource_id=invoice_id,
                ip_address=ip,
                details=f"File uploaded: {file.filename}",
            )
            return UploadedFileResponse.model_validate(new_file)

    async def get_file(self, invoice_id: str):
        file = await self.file_repo.get_by_invoice_id(invoice_id)
        if not file:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File not found")
        return file

    async def export_csv(self, filters: InvoiceFilters) -> str:
        import csv
        import io
        items = await self.invoice_repo.get_all_for_export(filters)
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "Invoice Number", "Serial Number", "Financial Year", "Bank Name",
            "Invoice Date", "Loan Requested Amount", "Loan Sanctioned Amount",
            "Has File", "Created At"
        ])
        for inv in items:
            writer.writerow([
                inv.invoice_number, inv.serial_number, inv.financial_year, inv.bank_name,
                inv.invoice_date.strftime("%Y-%m-%d") if inv.invoice_date else "",
                str(inv.loan_requested_amount), str(inv.loan_sanctioned_amount or ""),
                "Yes" if inv.uploaded_file else "No",
                inv.created_at.strftime("%Y-%m-%d %H:%M:%S") if inv.created_at else "",
            ])
        return output.getvalue()

    async def export_excel(self, filters: InvoiceFilters) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import io

        items = await self.invoice_repo.get_all_for_export(filters)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoices"

        headers = [
            "Invoice Number", "Serial Number", "Financial Year", "Bank Name",
            "Invoice Date", "Loan Requested (₹)", "Loan Sanctioned (₹)",
            "Has File", "Created At"
        ]
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row, inv in enumerate(items, 2):
            ws.cell(row=row, column=1, value=inv.invoice_number)
            ws.cell(row=row, column=2, value=inv.serial_number)
            ws.cell(row=row, column=3, value=inv.financial_year)
            ws.cell(row=row, column=4, value=inv.bank_name)
            ws.cell(row=row, column=5, value=inv.invoice_date.strftime("%Y-%m-%d") if inv.invoice_date else "")
            ws.cell(row=row, column=6, value=float(inv.loan_requested_amount))
            ws.cell(row=row, column=7, value=float(inv.loan_sanctioned_amount) if inv.loan_sanctioned_amount else 0)
            ws.cell(row=row, column=8, value="Yes" if inv.uploaded_file else "No")
            ws.cell(row=row, column=9, value=inv.created_at.strftime("%Y-%m-%d %H:%M:%S") if inv.created_at else "")

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
